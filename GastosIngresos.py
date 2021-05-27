import pandas as pd, os
from openpyxl import load_workbook 
from tkinter import *
from tkinter import messagebox
import pyautogui as pg

# IMPORTANTE RECORDAR QUE LOS ARCHIVOS "GastosIngresos.xlsx" y "monedas.png" DEBEN ESTAR GUARDADOS EN LA self.raiz DE LA PARTIDION D:\
# POR QUE EL PROGRAMA BUSCA ESTOS ARCHIVOS EN ESE PATH, TOMAR ESTO EN CUENTA A LA HORA DE DESEAR UTILIZAR EL PROGRAMA EJECUTABLE EN
# OTRA COMPUTADORA, TAMBIEN RECORDAR QUE COMO EL PROGRAMA FUE COMPILADO EN WINDOS SOLO SE PODRA EJECUTAR EN ESTE SISTEMA OPERATIVO 

#-----PARA HACER QUE SE MUESTREN TODAS LAS COLUMNAS DEl DATAFRAME EN EL WIDGET TEXT---------------------------------------------------
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

#-------------------- Creacion de la segunda ventana con sus funciones y widgets--------------------
class v2():

    def __init__(self):

        self.ventana = Tk()
        self.ventana.resizable(0, 0)
        self.ventana.iconbitmap('dollarsign.ico')
        self.ventana.geometry("1200x550")
        self.ventana.configure(bg = "beige")
        self.ventana.title("****** Reporte general de todos los meses ******")

        #Cargamos nuevamente el archivo pero esta vez usando pandas para crear un dataframe y pasarlo al widget text
        self.archivo = pd.ExcelFile('GastosIngresos.xlsx')
        self.df = self.archivo.parse("Principal")
        self.tabla = Text(self.ventana, padx=5,pady=5)
        self.tabla.configure(bg="beige", width=200, height=10,borderwidth=5, relief="ridge", fg="blue")
        self.tabla.insert(INSERT, str(self.df))
        self.tabla.pack()  

        self.delete = Label(self.ventana, text="Si desea puede eliminar registros de uno o todos los meses\n dando click sobre los botones correspondientes al Mes que se presentan a continuacion:", font=("arial"), fg="red", bg="beige")
        self.delete.pack(side= TOP, padx=20, pady=20)

        self.Otroframe = Frame(self.ventana, width="380", height="390")
        self.Otroframe.grid_propagate(False)
        self.Otroframe.config(bg="beige")
        self.Otroframe.pack(side=TOP)    
        
        #-------------------------------------------------------------------------------------------------
        #Creamos los botones que borraran los registros de cada mes segun el boton presionado, cuando se
        #hace el llamado de la funcion se hace a una lambda para poder pasar por parametro la letra que 
        #corresponde a la columna de dicho mes en el archivo excell
        self.EneroMes = Button(self.Otroframe, text="Enero", command=lambda: self.borrar_Mes('B'), width=10, height=1)
        self.EneroMes.grid(row=11, column=17, padx=5, pady=5)
        self.FebreroMes = Button(self.Otroframe, text="Febrero", command=lambda: self.borrar_Mes('C'), width=10, height=1)
        self.FebreroMes.grid(row=12, column=17, padx=5, pady=5)
        self.MarzoMes = Button(self.Otroframe, text="Marzo", command=lambda: self.borrar_Mes('D'), width=10, height=1)
        self.MarzoMes.grid(row=13, column=17, padx=5, pady=5)
        self.AbrilMes = Button(self.Otroframe, text="Abril", command=lambda: self.borrar_Mes('E'), width=10, height=1)
        self.AbrilMes.grid(row=14, column=17, padx=5, pady=5)
        self.MayoMes = Button(self.Otroframe, text="Mayo", command=lambda: self.borrar_Mes('F'), width=10, height=1)
        self.MayoMes.grid(row=11, column=19, padx=5, pady=5)
        self.JunioMes = Button(self.Otroframe, text="Junio", command=lambda: self.borrar_Mes('G'), width=10, height=1)
        self.JunioMes.grid(row=12, column=19, padx=5, pady=5)
        self.JulioMes = Button(self.Otroframe, text="Julio", command=lambda: self.borrar_Mes('H'), width=10, height=1)
        self.JulioMes.grid(row=13, column=19, padx=5, pady=5)
        self.AgostoMes = Button(self.Otroframe, text="Agosto", command=lambda: self.borrar_Mes('I'), width=10, height=1)
        self.AgostoMes.grid(row=14, column=19, padx=5, pady=5)
        self.SeptiembreMes = Button(self.Otroframe, text="Septiembre", command=lambda: self.borrar_Mes('J'), width=10, height=1)
        self.SeptiembreMes.grid(row=11, column=21, padx=5, pady=5)
        self.OctubreMes = Button(self.Otroframe, text="Octubre", command=lambda: self.borrar_Mes('K'), width=10, height=1)
        self.OctubreMes.grid(row=12, column=21, padx=5, pady=5)
        self.NoviembreMes = Button(self.Otroframe, text="Noviembre", command=lambda: self.borrar_Mes('L'), width=10, height=1)
        self.NoviembreMes.grid(row=13, column=21, padx=5, pady=5)
        self.DiciembreMes = Button(self.Otroframe, text="Diciembre", command=lambda: self.borrar_Mes('M'), width=10, height=1)
        self.DiciembreMes.grid(row=14, column=21, padx=5, pady=5)
        self.BorraTodo = Button(self.Otroframe, text="Borrar todos los registros", command=self.EliTodoMes, width=20, height=1)
        self.BorraTodo.grid(row=16, column=18, padx=5, pady=5, columnspan=3)
        
        self.ventana.mainloop()

    def borrar_Mes(self,columna):
        
        self.archivo2 = load_workbook('GastosIngresos.xlsx')  # cargar el archivo en la variable archivo2
        self.hoja1 = self.archivo2["Principal"] # Asignar los valores de la hoja "Principal" a la variable hoja1

        #Construimos la columna para modificar sus valores, como parametro se recibe la letra de la
        #columna y se le concatena el número, para ello nos apoyamos de un for y creamos una lisa con 
        #las letras y su correspondiente número para que puedan ser ubicadas en el archivo
        self.n_col = [columna + str(i) for i in range(2,11)]

        self.hoja1[self.n_col[0]] = float(0) 
        self.hoja1[self.n_col[1]] = float(0)
        self.hoja1[self.n_col[2]] = float(0)
        self.hoja1[self.n_col[3]] = float(0)
        self.hoja1[self.n_col[4]] = float(0)
        self.hoja1[self.n_col[5]] = float(0)
        self.hoja1[self.n_col[6]] = float(0)
        self.hoja1[self.n_col[7]] = float(0)
        self.hoja1[self.n_col[8]] = float(0)

        self.archivo2.save('GastosIngresos.xlsx')
        
        self.tabla.delete(1.0,END)
        self.archivo = pd.ExcelFile('GastosIngresos.xlsx')
        self.df = self.archivo.parse("Principal")
        self.tabla.insert("insert", str(self.df))

    def EliTodoMes(self):

        #Pedimos que se ingrese el nonmbre que tendra la hoja de respaldo
        self.nombre_hoja = pg.prompt('Por favor ingresa el nombre con el que desea respaldar la información')

        self.book = load_workbook('GastosIngresos.xlsx')
        self.hoja_principal = self.book['Principal']

        self.book.copy_worksheet(self.hoja_principal)

        self.hoja_respaldo = self.book['Principal Copy']
        self.hoja_respaldo.title = self.nombre_hoja

        self.celdas = self.hoja_principal['B2':'M10']

        for filas in self.celdas:
            for celda in filas:
                celda.value = float(0)

        self.book.save('GastosIngresos.xlsx')
                
        #Finalmente volvemos a presentar la informacion ya con los datos iguales a 0
        self.tabla.delete(1.0,END)
        self.archivo = pd.ExcelFile('GastosIngresos.xlsx')
        self.df = self.archivo.parse("Principal")
        self.tabla.insert("insert", str(self.df))

class app():

    def __init__(self):

        self.raiz = Tk()
        self.raiz.resizable(0, 0)
        self.raiz.iconbitmap('dollarsign.ico')
        self.raiz.geometry("840x580")
        self.raiz.configure(bg = "beige")
        self.raiz.title("****** Bienvenido al calculo de Balance del Mes ******")

        #---------------------------- creacion de el FRAME--------------------------------------------------------

        self.El_frame = Frame(self.raiz, width="380", height="570")
        self.El_frame.grid_propagate(False)
        self.El_frame.config(bg="beige")
        self.El_frame.pack(side=LEFT, expand=1)

        self.nwframe = Frame(self.raiz, width="520", height="570")
        self.nwframe.grid_propagate(False)
        self.nwframe.config(bg="beige")
        self.nwframe.pack(side=RIGHT)

        self.imagen = PhotoImage(file="D:/monedas.png")
        self.ima = Label(self.nwframe, image= self.imagen)
        self.ima.place(relx=0.2, rely=0.2)

        #---------------------------- creacion de los label que van al lado de los cuadro de texto-----------------------------------------------

        self.datos = Label(self.El_frame, text="Ingrese los siguientes datos:", bg="beige", font="Helvetica 16 bold")
        self.datos.grid(row=3, column=9,padx=10, pady=10, columnspan=2)

        self.MesEvaluado = Label(self.El_frame, text="Ingrese el Mes a Evaluar", bg="beige", font="Arial")
        self.MesEvaluado.grid(row=4, column=9,padx=10, pady=10)

        self.sueldo1 = Label(self.El_frame, text="Ingrese el sueldo", bg="beige", font="Arial")
        self.sueldo1.grid(row=5, column=9, padx=10, pady=10)

        self.IngExtras1 = Label(self.El_frame, text="Ingresos extras", bg="beige", font="Arial")
        self.IngExtras1.grid(row=6, column=9, padx=10, pady=10)

        self.GastHogar1 = Label(self.El_frame, text="Gastos Hogar", bg="beige", font="Arial")
        self.GastHogar1.grid(row=7, column=9, padx=10, pady=10)

        self.Vehiculo1 = Label(self.El_frame, text="Gastos Vehiculo", bg="beige", font="Arial")
        self.Vehiculo1.grid(row=8, column=9, padx=10, pady=10)

        self.Ocio1 = Label(self.El_frame, text="Gastos de Ocio", bg="beige", font="Arial")
        self.Ocio1.grid(row=9, column=9, padx=10, pady=10)

        self.OtrosGast1 = Label(self.El_frame, text="Otros Gastos", bg="beige", font="Arial")
        self.OtrosGast1.grid(row=10, column=9, padx=10, pady=10)


        #---------------------------- creacion de los cuardo de texto-----------------------------------------------

        self.GuardaMes = StringVar()

        self.MesEva=Entry(self.El_frame, textvariable=self.GuardaMes)
        self.MesEva.grid(row=4, column=10, padx=10, pady=10)

        self.suld = StringVar() # Variabla a ser asociada con el Entry de nombre sueldo

        self.sueldo=Entry(self.El_frame, textvariable=self.suld)
        self.sueldo.grid(row=5, column=10, padx=10, pady=10)

        self.Ingre = StringVar()

        self.IngExtras=Entry(self.El_frame, textvariable=self.Ingre)
        self.IngExtras.grid(row=6, column=10, padx=10, pady=10)

        self.Gast = StringVar()

        self.GastHogar=Entry(self.El_frame, textvariable=self.Gast)
        self.GastHogar.grid(row=7, column=10, padx=10, pady=10)

        self.Veh = StringVar()

        self.Vehiculo=Entry(self.El_frame, textvariable=self.Veh)
        self.Vehiculo.grid(row=8, column=10, padx=10, pady=10)

        self.Oci = StringVar()

        self.Ocio=Entry(self.El_frame, textvariable=self.Oci)
        self.Ocio.grid(row=9, column=10, padx=10, pady=10)

        self.OtroGas = StringVar()

        self.OtrosGast=Entry(self.El_frame, textvariable=self.OtroGas)
        self.OtrosGast.grid(row=10, column=10,padx=10, pady=10)

        #----------------------------- creacion de los botones ---------------------------------------------------
        self.salir = Button(self.El_frame, text='Salir', command=self.raiz.destroy, width=20, height=2, activebackground="#E6F0FF")
        self.salir.grid(row=15, column=10, padx=10, pady=10, columnspan=2)

        self.calcular = Button(self.El_frame, text="Calcular", width=10, command=self.calc)
        self.calcular.grid(row=11, column=10, padx=10, pady=10, columnspan=2)

        self.limpiar = Button(self.El_frame, text="Limpiar datos", command=self.borrar)
        self.limpiar.grid(row=13, column=10, padx=10, pady=10, columnspan=2)

        self.informe = Button(self.El_frame, text="Ver Informe Completo", width=20, command=v2)
        self.informe.grid(row=14, column=10, padx=10, pady=10)

        #----- main loop para que no se cierre la app ------------------------
        self.raiz.mainloop()
        
    #--------------------------- creacion de funciones para los botones ---------------------------------------

    def borrar(self):

        self.GuardaMes.set("")
        self.suld.set("")
        self.Ingre.set("")
        self.Gast.set("")
        self.Veh.set("")
        self.Oci.set("")
        self.OtroGas.set("")
        self.res1.config(text='')

    def calc(self):

        self.archivo2 = load_workbook('GastosIngresos.xlsx')  # cargar el archivo en la variable archivo2 
        self.hoja1 = self.archivo2["Principal"] # Asignar los valores de la hoja "Principal" a la variable hoja1

        self.Lista_Meses = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
        'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'] 

        self.sul1 = self.suld.get()
        self.Ingre1 = self.Ingre.get()
        self.Gas1 = self.Gast.get()
        self.Oci1 = self.Oci.get()
        self.OtroGas1 = self.OtroGas.get()
        self.Veh1 = self.Veh.get() 

        self.dic_meses = {
            'ENERO':['B'+str(j) for j in range(2,11)],
            'FEBRERO':['C'+str(j) for j in range(2,11)],
            'MARZO':['D'+str(j) for j in range(2,11)],
            'ABRIL':['E'+str(j) for j in range(2,11)],
            'MAYO':['F'+str(j) for j in range(2,11)],
            'JUNIO':['G'+str(j) for j in range(2,11)],
            'JULIO':['H'+str(j) for j in range(2,11)],
            'AGOSTO':['I'+str(j) for j in range(2,11)],
            'SEPTIEMBRE':['J'+str(j) for j in range(2,11)],
            'OCTUBRE':['K'+str(j) for j in range(2,11)],
            'NOVIEMBRE':['L'+str(j) for j in range(2,11)],
            'DICIEMBRE':['M'+str(j) for j in range(2,11)]
        }

        self.mes_ingresado = str(self.GuardaMes.get()).upper()

        self.Nombre_cajas = [(self.suld,self.sueldo), (self.Ingre,self.IngExtras),
        (self.Gast,self.GastHogar),(self.Veh,self.Vehiculo),(self.Oci,self.Ocio),
        (self.OtroGas,self.OtrosGast)] 

        if self.mes_ingresado in self.Lista_Meses:
            
            cf2,cf3,cf4,cf5,cf6,cf7,cf8,cf9,cf10 = self.dic_meses[self.mes_ingresado]

            #Con el siguiente ciclo for preguntamos si el contenido de las cajas de texto estan vacios o si NO es un dato 
            # numerico, Nombre_Cajas es una lista compuesta por tuplas con 2 elementos cada tupla por eso usamos h y j  
            for h,j in self.Nombre_cajas:    
                    k = h.get().replace(',','').replace('.','')
                    if h.get().replace(',','') == '' or not k.isnumeric():
                        h.set('')
                        j.insert(0,'0')
                
            self.hoja1[cf2] = float(self.suld.get().replace(',','')) + self.hoja1[cf2].value
            self.hoja1[cf3] = float(self.Ingre.get().replace(',','')) + self.hoja1[cf3].value
            self.hoja1[cf5] = float(self.Gast.get().replace(',','')) + self.hoja1[cf5].value
            self.hoja1[cf6] = float(self.Veh.get().replace(',','')) + self.hoja1[cf6].value
            self.hoja1[cf7] = float(self.Oci.get().replace(',','')) + self.hoja1[cf7].value
            self.hoja1[cf8] = float(self.OtroGas.get().replace(',','')) + self.hoja1[cf8].value   
            self.TotalIng =  (self.hoja1[cf2].value) + (self.hoja1[cf3].value)
            self.hoja1[cf4] = self.TotalIng
            self.TotalGas = (self.hoja1[cf5].value) + (self.hoja1[cf6].value) + (self.hoja1[cf7].value) + (self.hoja1[cf8].value)
            self.hoja1[cf9] = self.TotalGas
            self.Balance = self.TotalIng - self.TotalGas
            self.hoja1[cf10] = self.Balance
            self.archivo2.save('GastosIngresos.xlsx')
            
            #Para formatear el resultado de los calculos de TotalIng,TotalGas y Balance hacemos uso de las
            #f strings y especificamos la coma (,) para los miles y .2f para que solo nos muestre 2 decimales
            self.texto_resultados =f'El ingreso total es de: ${self.TotalIng:,.2f} \n\n El gasto total es de: ${self.TotalGas:,.2f} \n\n El ahorro fue de: ${self.Balance:,.2f}'
            
            self.res1 = Label(self.nwframe, width=30, bg="beige", text=self.texto_resultados)
            self.res1.config(font="Arial 11 bold")
            self.res1.place(relx=0.4, rely=0.7, anchor=CENTER)

        else:
            messagebox.showwarning("Advertencia", "Ingrese un nombre de Mes valido (Enero, Febrero..)")

if __name__ == "__main__":
    app()
